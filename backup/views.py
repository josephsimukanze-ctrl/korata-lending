# backup/views.py
import os
import json
import zipfile
import subprocess
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, FileResponse
from django.core.management import call_command
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from django.db import connection
from .models import BackupLog
from users.models import CustomUser
from clients.models import Client
from loans.models import Loan, Payment
from collateral.models import Collateral
import logging

logger = logging.getLogger(__name__)

def is_admin_or_ceo(user):
    return user.is_superuser or (hasattr(user, 'role') and user.role in ['ceo', 'admin'])

@login_required
@user_passes_test(is_admin_or_ceo)
def backup_dashboard(request):
    """Main backup dashboard"""
    backups = BackupLog.objects.all()[:20]
    
    context = {
        'backups': backups,
        'total_backups': BackupLog.objects.count(),
        'total_size': sum(b.file_size for b in backups),
        'title': 'Backup & Restore Dashboard'
    }
    return render(request, 'backup/dashboard.html', context)

@login_required
@user_passes_test(is_admin_or_ceo)
def backup_database(request):
    """Create database backup"""
    if request.method == 'POST':
        try:
            # Create backup log
            backup_log = BackupLog.objects.create(
                operation_type='database',
                status='running',
                created_by=request.user
            )
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"korata_db_backup_{timestamp}.json"
            backup_path = os.path.join(settings.BASE_DIR, 'backups', filename)
            
            # Create backup directory if it doesn't exist
            os.makedirs(os.path.dirname(backup_path), exist_ok=True)
            
            # Backup using dumpdata
            with open(backup_path, 'w') as f:
                call_command('dumpdata', indent=2, stdout=f)
            
            # Get file size
            file_size = os.path.getsize(backup_path)
            
            # Update backup log
            backup_log.status = 'completed'
            backup_log.filename = filename
            backup_log.file_size = file_size
            backup_log.completed_at = datetime.now()
            backup_log.save()
            
            messages.success(request, f'Database backup created successfully: {filename}')
            
        except Exception as e:
            logger.error(f"Backup failed: {e}")
            backup_log.status = 'failed'
            backup_log.error_message = str(e)
            backup_log.save()
            messages.error(request, f'Backup failed: {str(e)}')
        
        return redirect('backup:dashboard')
    
    return render(request, 'backup/backup_database.html')

@login_required
@user_passes_test(is_admin_or_ceo)
def backup_media(request):
    """Create media files backup"""
    if request.method == 'POST':
        try:
            backup_log = BackupLog.objects.create(
                operation_type='media',
                status='running',
                created_by=request.user
            )
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"korata_media_backup_{timestamp}.zip"
            backup_path = os.path.join(settings.BASE_DIR, 'backups', filename)
            
            os.makedirs(os.path.dirname(backup_path), exist_ok=True)
            
            # Create zip file of media directory
            with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root, dirs, files in os.walk(settings.MEDIA_ROOT):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, settings.MEDIA_ROOT)
                        zipf.write(file_path, arcname)
            
            file_size = os.path.getsize(backup_path)
            
            backup_log.status = 'completed'
            backup_log.filename = filename
            backup_log.file_size = file_size
            backup_log.completed_at = datetime.now()
            backup_log.save()
            
            messages.success(request, f'Media backup created successfully: {filename}')
            
        except Exception as e:
            logger.error(f"Media backup failed: {e}")
            backup_log.status = 'failed'
            backup_log.error_message = str(e)
            backup_log.save()
            messages.error(request, f'Media backup failed: {str(e)}')
        
        return redirect('backup:dashboard')
    
    return render(request, 'backup/backup_media.html')

@login_required
@user_passes_test(is_admin_or_ceo)
def full_backup(request):
    """Create full system backup"""
    if request.method == 'POST':
        try:
            backup_log = BackupLog.objects.create(
                operation_type='full',
                status='running',
                created_by=request.user
            )
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"korata_full_backup_{timestamp}.zip"
            backup_path = os.path.join(settings.BASE_DIR, 'backups', filename)
            
            os.makedirs(os.path.dirname(backup_path), exist_ok=True)
            
            # Create zip file containing both database and media
            with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Backup database to JSON
                db_backup = f"database_backup_{timestamp}.json"
                with open(db_backup, 'w') as f:
                    call_command('dumpdata', indent=2, stdout=f)
                zipf.write(db_backup, db_backup)
                os.remove(db_backup)
                
                # Add media files
                for root, dirs, files in os.walk(settings.MEDIA_ROOT):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.join('media', os.path.relpath(file_path, settings.MEDIA_ROOT))
                        zipf.write(file_path, arcname)
            
            file_size = os.path.getsize(backup_path)
            
            backup_log.status = 'completed'
            backup_log.filename = filename
            backup_log.file_size = file_size
            backup_log.completed_at = datetime.now()
            backup_log.save()
            
            messages.success(request, f'Full backup created successfully: {filename}')
            
        except Exception as e:
            logger.error(f"Full backup failed: {e}")
            backup_log.status = 'failed'
            backup_log.error_message = str(e)
            backup_log.save()
            messages.error(request, f'Full backup failed: {str(e)}')
        
        return redirect('backup:dashboard')
    
    return render(request, 'backup/full_backup.html')

@login_required
@user_passes_test(is_admin_or_ceo)
def restore_backup(request, backup_id):
    """Restore from backup"""
    backup = get_object_or_404(BackupLog, id=backup_id)
    
    if request.method == 'POST':
        try:
            restore_log = BackupLog.objects.create(
                operation_type='restore',
                status='running',
                created_by=request.user,
                notes=f"Restoring from backup: {backup.filename}"
            )
            
            backup_path = os.path.join(settings.BASE_DIR, 'backups', backup.filename)
            
            if not os.path.exists(backup_path):
                raise Exception(f"Backup file not found: {backup_path}")
            
            if backup.operation_type == 'database':
                # Restore database from JSON
                with open(backup_path, 'r') as f:
                    call_command('loaddata', backup_path, stdout=f)
            
            elif backup.operation_type == 'full':
                # Extract full backup
                with zipfile.ZipFile(backup_path, 'r') as zipf:
                    # Restore database
                    db_files = [f for f in zipf.namelist() if f.startswith('database_backup_')]
                    if db_files:
                        zipf.extract(db_files[0])
                        with open(db_files[0], 'r') as f:
                            call_command('loaddata', db_files[0], stdout=f)
                        os.remove(db_files[0])
                    
                    # Restore media files
                    for file in zipf.namelist():
                        if file.startswith('media/'):
                            zipf.extract(file, settings.BASE_DIR)
            
            restore_log.status = 'completed'
            restore_log.completed_at = datetime.now()
            restore_log.save()
            
            messages.success(request, 'Backup restored successfully!')
            
        except Exception as e:
            logger.error(f"Restore failed: {e}")
            restore_log.status = 'failed'
            restore_log.error_message = str(e)
            restore_log.save()
            messages.error(request, f'Restore failed: {str(e)}')
        
        return redirect('backup:dashboard')
    
    return render(request, 'backup/restore_confirm.html', {'backup': backup})

@login_required
@user_passes_test(is_admin_or_ceo)
def download_backup(request, backup_id):
    """Download backup file"""
    backup = get_object_or_404(BackupLog, id=backup_id)
    backup_path = os.path.join(settings.BASE_DIR, 'backups', backup.filename)
    
    if os.path.exists(backup_path):
        response = FileResponse(open(backup_path, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{backup.filename}"'
        return response
    else:
        messages.error(request, 'Backup file not found')
        return redirect('backup:dashboard')

@login_required
@user_passes_test(is_admin_or_ceo)
def delete_backup(request, backup_id):
    """Delete backup file"""
    backup = get_object_or_404(BackupLog, id=backup_id)
    
    if request.method == 'POST':
        backup_path = os.path.join(settings.BASE_DIR, 'backups', backup.filename)
        if os.path.exists(backup_path):
            os.remove(backup_path)
        backup.delete()
        messages.success(request, 'Backup deleted successfully')
        return redirect('backup:dashboard')
    
    return render(request, 'backup/delete_confirm.html', {'backup': backup})

@login_required
@user_passes_test(is_admin_or_ceo)
def export_data(request):
    """Export data in various formats"""
    export_format = request.GET.get('format', 'json')
    model_name = request.GET.get('model', 'all')
    
    # Define models to export
    models_map = {
        'users': CustomUser,
        'clients': Client,
        'loans': Loan,
        'payments': Payment,
        'collateral': Collateral,
    }
    
    data = {}
    
    if model_name == 'all':
        for name, model in models_map.items():
            queryset = model.objects.all().values()
            data[name] = list(queryset)
    elif model_name in models_map:
        queryset = models_map[model_name].objects.all().values()
        data[model_name] = list(queryset)
    else:
        return JsonResponse({'error': 'Invalid model'}, status=400)
    
    # Create response based on format
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"korata_export_{model_name}_{timestamp}.{export_format}"
    
    if export_format == 'json':
        response = HttpResponse(json.dumps(data, indent=2, default=str), content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    elif export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        import csv
        writer = csv.writer(response)
        
        # Write headers and data
        for model_name, model_data in data.items():
            writer.writerow([f'--- {model_name.upper()} ---'])
            if model_data:
                writer.writerow(model_data[0].keys())
                for row in model_data:
                    writer.writerow(row.values())
            writer.writerow([])
        
        return response
    
    elif export_format == 'excel':
        from openpyxl import Workbook
        wb = Workbook()
        
        for model_name, model_data in data.items():
            if model_data:
                ws = wb.create_sheet(title=model_name[:31])
                headers = list(model_data[0].keys())
                ws.append(headers)
                for row in model_data:
                    ws.append(list(row.values()))
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    return JsonResponse({'error': 'Invalid format'}, status=400)

# backup/views.py - Improved import_data function

@login_required
@user_passes_test(is_admin_or_ceo)
def import_data(request):
    """Import data from file with better error handling"""
    if request.method == 'POST' and request.FILES.get('import_file'):
        import_file = request.FILES['import_file']
        import_mode = request.POST.get('mode', 'merge')
        
        # Create log entry
        import_log = BackupLog.objects.create(
            operation_type='import',
            status='running',
            created_by=request.user,
            filename=import_file.name
        )
        
        try:
            # Read file content
            content = import_file.read()
            file_extension = import_file.name.split('.')[-1].lower()
            
            imported_count = 0
            updated_count = 0
            skipped_count = 0
            errors = []
            
            if file_extension == 'json':
                data = json.loads(content.decode('utf-8'))
                
                # Define model mapping
                model_map = {
                    'users': CustomUser,
                    'clients': Client,
                    'loans': Loan,
                    'payments': Payment,
                    'collateral': Collateral,
                }
                
                # Handle Django dumpdata format
                if isinstance(data, list):
                    for idx, item in enumerate(data):
                        try:
                            if 'model' in item and 'fields' in item:
                                model_path = item.get('model')
                                model_name = model_path.split('.')[-1] if '.' in model_path else model_path
                                
                                if model_name in model_map:
                                    model = model_map[model_name]
                                    obj_id = item.get('pk')
                                    fields = item.get('fields', {})
                                    
                                    # Clean up fields - remove relations that might cause issues
                                    safe_fields = {}
                                    for key, value in fields.items():
                                        # Skip related objects that might not exist
                                        if key.endswith('_id') or key in ['password', 'is_superuser', 'is_staff']:
                                            continue
                                        safe_fields[key] = value
                                    
                                    try:
                                        if import_mode == 'replace':
                                            if obj_id:
                                                model.objects.filter(id=obj_id).delete()
                                            obj = model.objects.create(**safe_fields)
                                            imported_count += 1
                                        elif import_mode == 'skip':
                                            if obj_id and not model.objects.filter(id=obj_id).exists():
                                                obj = model.objects.create(**safe_fields)
                                                imported_count += 1
                                            else:
                                                skipped_count += 1
                                        else:  # merge
                                            obj, created = model.objects.update_or_create(
                                                id=obj_id,
                                                defaults=safe_fields
                                            )
                                            if created:
                                                imported_count += 1
                                            else:
                                                updated_count += 1
                                    except Exception as e:
                                        errors.append(f"Item {idx} - {model_name} ID {obj_id}: {str(e)}")
                                else:
                                    errors.append(f"Unknown model: {model_path}")
                            else:
                                errors.append(f"Invalid item format at index {idx}")
                        except Exception as e:
                            errors.append(f"Error processing item {idx}: {str(e)}")
                
                elif isinstance(data, dict):
                    # Handle keyed data format
                    for model_name, model_data in data.items():
                        if model_name in model_map:
                            model = model_map[model_name]
                            
                            if isinstance(model_data, list):
                                for idx, item in enumerate(model_data):
                                    try:
                                        if import_mode == 'replace':
                                            model.objects.all().delete()
                                            obj = model.objects.create(**item)
                                            imported_count += 1
                                        else:
                                            obj_id = item.get('id')
                                            if obj_id:
                                                obj, created = model.objects.update_or_create(
                                                    id=obj_id,
                                                    defaults=item
                                                )
                                                if created:
                                                    imported_count += 1
                                                else:
                                                    updated_count += 1
                                            else:
                                                obj = model.objects.create(**item)
                                                imported_count += 1
                                    except Exception as e:
                                        errors.append(f"Error in {model_name} item {idx}: {str(e)}")
                            else:
                                errors.append(f"Invalid data format for {model_name}")
                        else:
                            errors.append(f"Unknown model: {model_name}")
                else:
                    errors.append("Unsupported JSON format")
                    
            elif file_extension == 'csv':
                import csv
                import io
                
                csv_file = io.StringIO(content.decode('utf-8'))
                reader = csv.DictReader(csv_file)
                
                # Get model from first row or filename
                model_name = request.POST.get('model', 'clients')
                
                if model_name in model_map:
                    model = model_map[model_name]
                    
                    for idx, row in enumerate(reader):
                        try:
                            # Clean up row data
                            cleaned_row = {}
                            for key, value in row.items():
                                if value and value != '':
                                    # Try to convert numeric values
                                    if value.isdigit():
                                        cleaned_row[key] = int(value)
                                    elif value.replace('.', '').isdigit():
                                        cleaned_row[key] = float(value)
                                    else:
                                        cleaned_row[key] = value
                            
                            obj_id = cleaned_row.get('id')
                            if import_mode == 'replace':
                                if obj_id:
                                    model.objects.filter(id=obj_id).delete()
                                obj = model.objects.create(**cleaned_row)
                                imported_count += 1
                            elif import_mode == 'skip':
                                if obj_id and not model.objects.filter(id=obj_id).exists():
                                    obj = model.objects.create(**cleaned_row)
                                    imported_count += 1
                                else:
                                    skipped_count += 1
                            else:  # merge
                                if obj_id:
                                    obj, created = model.objects.update_or_create(
                                        id=obj_id,
                                        defaults=cleaned_row
                                    )
                                    if created:
                                        imported_count += 1
                                    else:
                                        updated_count += 1
                                else:
                                    obj = model.objects.create(**cleaned_row)
                                    imported_count += 1
                        except Exception as e:
                            errors.append(f"CSV row {idx + 2}: {str(e)}")
                else:
                    errors.append(f"Unknown model: {model_name}")
                    
            elif file_extension == 'xlsx':
                import openpyxl
                import io
                
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(cell.value)
                
                model_name = request.POST.get('model', 'clients')
                
                if model_name in model_map:
                    model = model_map[model_name]
                    
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            row_data = {}
                            for idx, header in enumerate(headers):
                                if idx < len(row) and row[idx] is not None:
                                    value = row[idx]
                                    # Try to convert numeric values
                                    if isinstance(value, (int, float)):
                                        row_data[header] = value
                                    elif str(value).isdigit():
                                        row_data[header] = int(value)
                                    else:
                                        row_data[header] = str(value)
                            
                            obj_id = row_data.get('id')
                            if import_mode == 'replace':
                                if obj_id:
                                    model.objects.filter(id=obj_id).delete()
                                obj = model.objects.create(**row_data)
                                imported_count += 1
                            elif import_mode == 'skip':
                                if obj_id and not model.objects.filter(id=obj_id).exists():
                                    obj = model.objects.create(**row_data)
                                    imported_count += 1
                                else:
                                    skipped_count += 1
                            else:
                                if obj_id:
                                    obj, created = model.objects.update_or_create(
                                        id=obj_id,
                                        defaults=row_data
                                    )
                                    if created:
                                        imported_count += 1
                                    else:
                                        updated_count += 1
                                else:
                                    obj = model.objects.create(**row_data)
                                    imported_count += 1
                        except Exception as e:
                            errors.append(f"Excel row {row_idx}: {str(e)}")
                else:
                    errors.append(f"Unknown model: {model_name}")
            
            else:
                errors.append(f"Unsupported file format: {file_extension}")
            
            # Update log with results
            import_log.status = 'completed' if len(errors) == 0 else 'failed'
            import_log.completed_at = datetime.now()
            import_log.notes = f"Imported: {imported_count}, Updated: {updated_count}, Skipped: {skipped_count}, Errors: {len(errors)}"
            
            if errors:
                import_log.error_message = '\n'.join(errors[:10])  # Store first 10 errors
                messages.warning(request, f'Import completed with {len(errors)} error(s). {imported_count} imported, {updated_count} updated, {skipped_count} skipped.')
                logger.warning(f"Import errors: {errors}")
            else:
                messages.success(request, f'Successfully imported {imported_count} records and updated {updated_count} records!')
                
            import_log.save()
            
        except Exception as e:
            logger.error(f"Import failed: {str(e)}")
            import_log.status = 'failed'
            import_log.error_message = str(e)
            import_log.completed_at = datetime.now()
            import_log.save()
            messages.error(request, f'Import failed: {str(e)}')
        
        return redirect('backup:dashboard')
    
    return render(request, 'backup/import_data.html')
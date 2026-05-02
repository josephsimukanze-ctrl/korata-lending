# core/views.py - Add/Update these functions

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

from clients.models import Client
from loans.models import Loan, Payment
from collateral.models import Collateral, AssetType
from .models import Report

def is_admin_or_ceo(user):
    return user.is_superuser or (hasattr(user, 'role') and user.role in ['ceo', 'admin'])

@login_required
@user_passes_test(is_admin_or_ceo)
def report_index(request):
    """Main reports dashboard"""
    from loans.models import Loan, Payment
    from clients.models import Client
    from collateral.models import Collateral
    from django.db.models import Sum
    from django.utils import timezone
    from datetime import timedelta
    
    # Get statistics
    total_clients = Client.objects.count()
    active_clients = Client.objects.filter(status='active').count()
    total_loans = Loan.objects.count()
    active_loans = Loan.objects.filter(status='active').count()
    total_collections = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    
    # Monthly collections
    today = timezone.now().date()
    month_start = today.replace(day=1)
    monthly_collections = Payment.objects.filter(
        payment_date__date__gte=month_start
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Net profit (simplified)
    total_disbursed = Loan.objects.aggregate(total=Sum('principal'))['total'] or 0
    net_profit = total_collections - total_disbursed
    
    # Collateral value
    collateral_value = Collateral.objects.aggregate(total=Sum('estimated_value'))['total'] or 0
    
    # Recent reports
    recent_reports = Report.objects.all().order_by('-created_at')[:5]
    
    context = {
        'report_count': Report.objects.count(),
        'total_collections': total_collections,
        'total_loans': total_loans,
        'active_clients': active_clients,
        'monthly_collections': monthly_collections,
        'active_loans': active_loans,
        'total_clients': total_clients,
        'net_profit': net_profit,
        'collateral_value': collateral_value,
        'recent_reports': recent_reports,
    }
    
    # Check if print format is requested
    if request.GET.get('format') == 'print':
        return render(request, 'reports/index_print.html', context)
    
    return render(request, 'reports/index.html', context)

@login_required
@user_passes_test(is_admin_or_ceo)
def collections_report(request):
    """Collections report - payments collected over time"""
    
    # Get date range
    days = int(request.GET.get('days', 30))
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=days)
    
    # Get payments in date range
    payments = Payment.objects.filter(
        payment_date__date__gte=start_date,
        payment_date__date__lte=end_date
    ).select_related('loan__client', 'collected_by')
    
    # Summary statistics
    total_collected = payments.aggregate(total=Sum('amount'))['total'] or 0
    total_transactions = payments.count()
    average_payment = total_collected / total_transactions if total_transactions > 0 else 0
    
    # Group by day
    daily_data = []
    current_date = start_date
    while current_date <= end_date:
        day_payments = payments.filter(payment_date__date=current_date)
        daily_data.append({
            'date': current_date,
            'amount': day_payments.aggregate(total=Sum('amount'))['total'] or 0,
            'count': day_payments.count()
        })
        current_date += timedelta(days=1)
    
    # Group by payment method
    method_dict = {}
    for payment in payments:
        method_name = payment.payment_method if payment.payment_method else 'Cash'
        
        if hasattr(payment, 'get_payment_method_display'):
            method_name = payment.get_payment_method_display()
        
        if method_name not in method_dict:
            method_dict[method_name] = {'amount': 0, 'count': 0}
        method_dict[method_name]['amount'] += float(payment.amount)
        method_dict[method_name]['count'] += 1
    
    method_data = []
    for method_name, data in method_dict.items():
        method_data.append({
            'method': method_name,
            'amount': data['amount'],
            'count': data['count']
        })
    method_data.sort(key=lambda x: x['amount'], reverse=True)
    
    # Recent collections
    recent_collections = payments.order_by('-payment_date')[:10]
    
    context = {
        'title': 'Collections Report',
        'start_date': start_date,
        'end_date': end_date,
        'days': days,
        'total_collected': total_collected,
        'total_transactions': total_transactions,
        'average_payment': average_payment,
        'daily_data': daily_data,
        'method_data': method_data,
        'recent_collections': recent_collections,
        'report_date': timezone.now(),
    }
    
    # Check if print format is requested
    if request.GET.get('format') == 'print':
        return render(request, 'reports/collections_report_print.html', context)
    
    return render(request, 'reports/collections_report.html', context)

@login_required
@user_passes_test(is_admin_or_ceo)
def loans_report(request):
    """Loans report - all loan activity"""
    
    # Get filters
    status_filter = request.GET.get('status', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Base queryset
    loans = Loan.objects.all().select_related('client')
    
    # Apply filters
    if status_filter:
        loans = loans.filter(status=status_filter)
    
    if start_date:
        loans = loans.filter(created_at__date__gte=start_date)
    if end_date:
        loans = loans.filter(created_at__date__lte=end_date)
    
    # Statistics
    total_loans = loans.count()
    total_principal = loans.aggregate(total=Sum('principal'))['total'] or 0
    total_interest = loans.aggregate(total=Sum('total_interest'))['total'] or 0
    total_payback = loans.aggregate(total=Sum('total_payback'))['total'] or 0
    avg_loan_size = total_principal / total_loans if total_loans > 0 else 0
    
    # Status breakdown
    status_breakdown = []
    status_choices = [
        ('pending', 'Pending'),
        ('approved', 'Approved'),
        ('active', 'Active'),
        ('completed', 'Completed'),
        ('defaulted', 'Defaulted'),
        ('rejected', 'Rejected'),
    ]
    for status_code, status_name in status_choices:
        count = loans.filter(status=status_code).count()
        if count > 0:
            status_breakdown.append({
                'status': status_name,
                'count': count,
                'percentage': (count / total_loans * 100) if total_loans > 0 else 0
            })
    
    # Recent loans
    recent_loans = loans.order_by('-created_at')[:10]
    
    context = {
        'title': 'Loans Report',
        'total_loans': total_loans,
        'total_principal': total_principal,
        'total_interest': total_interest,
        'total_payback': total_payback,
        'avg_loan_size': avg_loan_size,
        'status_breakdown': status_breakdown,
        'recent_loans': recent_loans,
        'status_filter': status_filter,
        'start_date': start_date,
        'end_date': end_date,
        'report_date': timezone.now(),
    }
    
    # Check if print format is requested
    if request.GET.get('format') == 'print':
        return render(request, 'reports/loans_report_print.html', context)
    
    return render(request, 'reports/loans_report.html', context)

@login_required
@user_passes_test(is_admin_or_ceo)
def clients_report(request):
    """Clients report - client demographics and statistics"""
    
    # Get all clients
    clients = Client.objects.all().select_related('registered_by')
    
    # Statistics
    total_clients = clients.count()
    active_clients = clients.filter(status='active').count()
    pending_kyc = clients.filter(kyc_verified=False).count()
    blacklisted = clients.filter(status='blacklisted').count()
    
    # City breakdown
    city_data = []
    cities = clients.values('city').annotate(count=Count('id')).order_by('-count')[:10]
    for city in cities:
        city_data.append({
            'city': city['city'] or 'Not specified',
            'count': city['count']
        })
    
    # Employment breakdown
    employment_choices = [
        ('employed', 'Employed'),
        ('self_employed', 'Self Employed'),
        ('business', 'Business Owner'),
        ('unemployed', 'Unemployed'),
        ('retired', 'Retired'),
    ]
    employment_data = []
    for emp_code, emp_name in employment_choices:
        count = clients.filter(employment_status=emp_code).count()
        if count > 0:
            employment_data.append({
                'status': emp_name,
                'count': count
            })
    
    # Registration trend (last 30 days)
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    registration_data = []
    current_date = start_date
    while current_date <= end_date:
        count = clients.filter(registration_date__date=current_date).count()
        registration_data.append({
            'date': current_date,
            'count': count
        })
        current_date += timedelta(days=1)
    
    # Recent clients
    recent_clients = clients.order_by('-registration_date')[:10]
    
    context = {
        'title': 'Clients Report',
        'total_clients': total_clients,
        'active_clients': active_clients,
        'pending_kyc': pending_kyc,
        'blacklisted': blacklisted,
        'city_data': city_data,
        'employment_data': employment_data,
        'registration_data': registration_data,
        'recent_clients': recent_clients,
        'report_date': timezone.now(),
    }
    
    # Check if print format is requested
    if request.GET.get('format') == 'print':
        return render(request, 'reports/clients_report_print.html', context)
    
    return render(request, 'reports/clients_report.html', context)

@login_required
@user_passes_test(is_admin_or_ceo)
def profit_loss_report(request):
    """Profit & Loss report"""
    
    # Get date range
    period = request.GET.get('period', 'month')
    
    if period == 'month':
        start_date = timezone.now().date().replace(day=1)
        end_date = timezone.now().date()
    elif period == 'quarter':
        quarter = (timezone.now().month - 1) // 3
        start_date = timezone.now().date().replace(month=quarter*3+1, day=1)
        end_date = timezone.now().date()
    elif period == 'year':
        start_date = timezone.now().date().replace(month=1, day=1)
        end_date = timezone.now().date()
    else:
        start_date = request.GET.get('start_date', timezone.now().date().replace(day=1))
        end_date = request.GET.get('end_date', timezone.now().date())
    
    # Calculate revenue (interest earned)
    loans_in_period = Loan.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    )
    interest_earned = loans_in_period.aggregate(total=Sum('total_interest'))['total'] or 0
    
    # Calculate disbursements
    total_disbursed = loans_in_period.aggregate(total=Sum('principal'))['total'] or 0
    
    # Calculate collections
    payments_in_period = Payment.objects.filter(
        payment_date__date__gte=start_date,
        payment_date__date__lte=end_date
    )
    total_collected = payments_in_period.aggregate(total=Sum('amount'))['total'] or 0
    
    # Calculate profit
    net_profit = interest_earned - total_disbursed + total_collected
    
    # Monthly breakdown
    monthly_data = []
    current_date = start_date
    while current_date <= end_date:
        month_start = current_date.replace(day=1)
        if current_date.month == 12:
            month_end = current_date.replace(year=current_date.year+1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = current_date.replace(month=current_date.month+1, day=1) - timedelta(days=1)
        
        month_loans = Loan.objects.filter(created_at__date__gte=month_start, created_at__date__lte=month_end)
        month_payments = Payment.objects.filter(payment_date__date__gte=month_start, payment_date__date__lte=month_end)
        
        monthly_data.append({
            'month': month_start.strftime('%b %Y'),
            'disbursed': month_loans.aggregate(total=Sum('principal'))['total'] or 0,
            'collected': month_payments.aggregate(total=Sum('amount'))['total'] or 0,
            'interest': month_loans.aggregate(total=Sum('total_interest'))['total'] or 0,
        })
        current_date = month_end + timedelta(days=1)
    
    context = {
        'title': 'Profit & Loss Report',
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        'interest_earned': interest_earned,
        'total_disbursed': total_disbursed,
        'total_collected': total_collected,
        'net_profit': net_profit,
        'monthly_data': monthly_data,
        'report_date': timezone.now(),
    }
    
    # Check if print format is requested
    if request.GET.get('format') == 'print':
        return render(request, 'reports/profit_loss_report_print.html', context)
    
    return render(request, 'reports/profit_loss_report.html', context)

@login_required
@user_passes_test(is_admin_or_ceo)
def collateral_report(request):
    """Generate collateral report with analytics"""
    from collateral.models import Collateral, AssetType
    from django.db.models import Sum, Count
    
    # Get filter parameters
    status = request.GET.get('status', '')
    asset_type_value = request.GET.get('asset_type', '')
    client_id = request.GET.get('client', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Start with all collaterals
    collaterals = Collateral.objects.select_related('asset_type', 'client').all()
    
    # Apply status filter
    if status:
        collaterals = collaterals.filter(status=status)
    
    # Apply asset type filter
    if asset_type_value:
        collaterals = collaterals.filter(asset_type__name=asset_type_value)
    
    # Apply client filter
    if client_id and client_id.isdigit():
        collaterals = collaterals.filter(client_id=int(client_id))
    
    # Apply date filters
    if date_from:
        collaterals = collaterals.filter(created_at__date__gte=date_from)
    if date_to:
        collaterals = collaterals.filter(created_at__date__lte=date_to)
    
    # Statistics
    total_collateral = collaterals.count()
    total_value = collaterals.aggregate(total=Sum('estimated_value'))['total'] or 0
    
    # Status breakdown
    status_breakdown = collaterals.values('status').annotate(
        count=Count('id'),
        value=Sum('estimated_value')
    ).order_by('status')
    
    # Asset type breakdown
    asset_type_breakdown = collaterals.values('asset_type__name').annotate(
        count=Count('id'),
        value=Sum('estimated_value')
    ).order_by('-value')
    
    # Condition breakdown
    condition_breakdown = collaterals.values('condition').annotate(
        count=Count('id'),
        value=Sum('estimated_value')
    ).order_by('condition')
    
    # Recently added
    recent_collateral = collaterals.order_by('-created_at')[:10]
    
    # For filter dropdowns
    asset_type_choices = AssetType.objects.filter(is_active=True)
    status_choices = [
        ('available', 'Available'),
        ('pledged', 'Pledged'),
        ('seized', 'Seized'),
        ('sold', 'Sold'),
        ('released', 'Released'),
    ]
    
    context = {
        'collaterals': recent_collateral,
        'total_collateral': total_collateral,
        'total_value': total_value,
        'status_breakdown': status_breakdown,
        'asset_type_breakdown': asset_type_breakdown,
        'condition_breakdown': condition_breakdown,
        'status_filter': status,
        'asset_type_filter': asset_type_value,
        'client_filter': client_id,
        'date_from': date_from,
        'date_to': date_to,
        'asset_type_choices': asset_type_choices,
        'status_choices': status_choices,
        'report_date': timezone.now(),
    }
    
    # Check if print format is requested
    if request.GET.get('format') == 'print':
        return render(request, 'reports/collateral_report_print.html', context)
    
    return render(request, 'reports/collateral_report.html', context)

# ==================== PRINT FUNCTIONS ====================

@login_required
@user_passes_test(is_admin_or_ceo)
def print_report_direct(request, report_type):
    """Direct print function - opens print dialog"""
    from django.shortcuts import redirect
    
    # Get the current URL parameters
    query_string = request.GET.urlencode()
    
    # Redirect to the appropriate report with print format
    if report_type == 'collections':
        return redirect(f'/reports/collections/?{query_string}&format=print')
    elif report_type == 'loans':
        return redirect(f'/reports/loans/?{query_string}&format=print')
    elif report_type == 'clients':
        return redirect(f'/reports/clients/?{query_string}&format=print')
    elif report_type == 'profit-loss':
        return redirect(f'/reports/profit-loss/?{query_string}&format=print')
    elif report_type == 'collateral':
        return redirect(f'/reports/collateral/?{query_string}&format=print')
    else:
        return redirect('/reports/')

# ==================== EXPORT FUNCTIONS ====================

@login_required
@user_passes_test(is_admin_or_ceo)
def export_report_csv(request, report_type):
    """Export report as CSV"""
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    
    if report_type == 'collections':
        writer.writerow(['Date', 'Amount', 'Receipt #', 'Client', 'Method', 'Collected By'])
        payments = Payment.objects.all().order_by('-payment_date')
        for payment in payments:
            writer.writerow([
                payment.payment_date.strftime('%Y-%m-%d'),
                payment.amount,
                payment.receipt_number,
                payment.loan.client.full_name if payment.loan else '',
                payment.payment_method.name if hasattr(payment, 'payment_method') and payment.payment_method else 'Cash',
                payment.created_by.username if payment.created_by else ''
            ])
    
    elif report_type == 'loans':
        writer.writerow(['Loan ID', 'Client', 'Principal', 'Interest', 'Total Payback', 'Status', 'Created'])
        loans = Loan.objects.all()
        for loan in loans:
            writer.writerow([
                loan.loan_id,
                loan.client.full_name,
                loan.principal,
                loan.total_interest or 0,
                loan.total_payback or 0,
                loan.get_status_display(),
                loan.created_at.strftime('%Y-%m-%d')
            ])
    
    elif report_type == 'clients':
        writer.writerow(['Client ID', 'Name', 'NRC', 'Phone', 'City', 'Status', 'KYC', 'Registered'])
        clients = Client.objects.all()
        for client in clients:
            writer.writerow([
                client.client_id,
                client.full_name,
                client.nrc,
                client.phone_number,
                client.city,
                client.get_status_display(),
                'Verified' if client.kyc_verified else 'Pending',
                client.registration_date.strftime('%Y-%m-%d')
            ])
    
    return response

@login_required
@user_passes_test(is_admin_or_ceo)
def export_report_excel(request, report_type):
    """Export report as Excel"""
    
    wb = Workbook()
    ws = wb.active
    
    if report_type == 'collections':
        ws.title = "Collections Report"
        headers = ['Date', 'Amount', 'Receipt #', 'Client', 'Method', 'Collected By']
        ws.append(headers)
        
        payments = Payment.objects.all().order_by('-payment_date')
        for payment in payments:
            ws.append([
                payment.payment_date.strftime('%Y-%m-%d'),
                float(payment.amount),
                payment.receipt_number,
                payment.loan.client.full_name if payment.loan else '',
                payment.payment_method.name if hasattr(payment, 'payment_method') and payment.payment_method else 'Cash',
                payment.created_by.username if payment.created_by else ''
            ])
    
    elif report_type == 'loans':
        ws.title = "Loans Report"
        headers = ['Loan ID', 'Client', 'Principal', 'Interest', 'Total Payback', 'Status', 'Created']
        ws.append(headers)
        
        loans = Loan.objects.all()
        for loan in loans:
            ws.append([
                loan.loan_id,
                loan.client.full_name,
                float(loan.principal),
                float(loan.total_interest or 0),
                float(loan.total_payback or 0),
                loan.get_status_display(),
                loan.created_at.strftime('%Y-%m-%d')
            ])
    
    elif report_type == 'clients':
        ws.title = "Clients Report"
        headers = ['Client ID', 'Name', 'NRC', 'Phone', 'City', 'Status', 'KYC', 'Registered']
        ws.append(headers)
        
        clients = Client.objects.all()
        for client in clients:
            ws.append([
                client.client_id,
                client.full_name,
                client.nrc,
                client.phone_number,
                client.city,
                client.get_status_display(),
                'Verified' if client.kyc_verified else 'Pending',
                client.registration_date.strftime('%Y-%m-%d')
            ])
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response
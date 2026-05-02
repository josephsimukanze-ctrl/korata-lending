from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from .models import Collateral, AssetType, CollateralInspection, CollateralMovement
from clients.models import Client
import csv
import json

# ==================== PERMISSION FUNCTIONS ====================

# collateral/views.py - Add these permission functions at the top

def is_admin_or_ceo(user):
    """Check if user is superuser, CEO, or Admin"""
    return user.is_superuser or (hasattr(user, 'role') and user.role in ['ceo', 'admin'])

def can_manage_collateral(user):
    """Check if user can manage collateral - Collateral Officer, Admin, CEO, Superuser"""
    return user.is_superuser or (hasattr(user, 'role') and user.role in ['ceo', 'admin', 'collateral_officer'])

def can_manage_insurance(user):
    """Check if user can manage insurance - Collateral Officer, Admin, CEO, Superuser"""
    return user.is_superuser or (hasattr(user, 'role') and user.role in ['ceo', 'admin', 'collateral_officer'])

def is_loan_officer(user):
    """Check if user is loan officer or higher"""
    return user.is_superuser or (hasattr(user, 'role') and user.role in ['ceo', 'admin', 'loan_officer'])

def is_accountant(user):
    """Check if user is accountant or higher"""
    return user.is_superuser or (hasattr(user, 'role') and user.role in ['ceo', 'admin', 'accountant'])

def can_view_reports(user):
    """Check if user can view reports"""
    return user.is_superuser or (hasattr(user, 'role') and user.role in ['ceo', 'admin', 'accountant', 'auditor'])

# ==================== COLLATERAL VIEWS ====================

@login_required
@user_passes_test(can_manage_collateral)
def collateral_list(request):
    """List all collateral items"""
    collaterals = Collateral.objects.all().select_related('client', 'asset_type')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        collaterals = collaterals.filter(
            Q(collateral_id__icontains=search_query) |
            Q(title__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(client__first_name__icontains=search_query) |
            Q(client__last_name__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        collaterals = collaterals.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(collaterals, 20)
    page = request.GET.get('page', 1)
    collaterals_page = paginator.get_page(page)
    
    # Statistics
    stats = {
        'total': Collateral.objects.count(),
        'pledged': Collateral.objects.filter(status='pledged').count(),
        'seized': Collateral.objects.filter(status='seized').count(),
        'verified': Collateral.objects.filter(verification_status='verified').count(),
        'total_value': float(Collateral.objects.aggregate(total=Sum('estimated_value'))['total'] or 0),
    }
    
    context = {
        'collaterals': collaterals_page,
        'stats': stats,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    
    return render(request, 'collateral/collateral_list.html', context)


@login_required
@user_passes_test(can_manage_collateral)
def collateral_create(request):
    """Create new collateral - CUSTOM FORM, NOT ADMIN"""
    # Get data for dropdowns
    clients = Client.objects.filter(status='active') if hasattr(Client, 'is_active') else Client.objects.all()
    asset_types = AssetType.objects.filter(is_active=True)
    
    if request.method == 'POST':
        try:
            # Get form data
            client_id = request.POST.get('client')
            asset_type_id = request.POST.get('asset_type')
            title = request.POST.get('title')
            description = request.POST.get('description', '')
            make = request.POST.get('make', '')
            model = request.POST.get('model', '')
            year = request.POST.get('year') or None
            color = request.POST.get('color', '')
            serial_number = request.POST.get('serial_number')
            condition = request.POST.get('condition')
            estimated_value = request.POST.get('estimated_value')
            loan_to_value_ratio = request.POST.get('loan_to_value_ratio', 70)
            storage_location = request.POST.get('storage_location')
            storage_section = request.POST.get('storage_section', '')
            storage_shelf = request.POST.get('storage_shelf', '')
            is_insured = request.POST.get('is_insured') == 'on'
            insurance_provider = request.POST.get('insurance_provider', '')
            insurance_policy_number = request.POST.get('insurance_policy_number', '')
            insurance_value = request.POST.get('insurance_value') or None
            insurance_expiry_date = request.POST.get('insurance_expiry_date') or None
            notes = request.POST.get('notes', '')
            tags = request.POST.get('tags', '')
            
            # Validate required fields
            if not all([client_id, asset_type_id, title, serial_number, estimated_value, storage_location]):
                messages.error(request, 'Please fill in all required fields.')
                return redirect('collateral:create')
            
            # Create collateral
            collateral = Collateral.objects.create(
                client_id=client_id,
                asset_type_id=asset_type_id,
                title=title,
                description=description,
                make=make,
                model=model,
                year=year,
                color=color,
                serial_number=serial_number,
                condition=condition,
                estimated_value=estimated_value,
                loan_to_value_ratio=loan_to_value_ratio,
                storage_location=storage_location,
                storage_section=storage_section,
                storage_shelf=storage_shelf,
                is_insured=is_insured,
                insurance_provider=insurance_provider,
                insurance_policy_number=insurance_policy_number,
                insurance_value=insurance_value,
                insurance_expiry_date=insurance_expiry_date,
                notes=notes,
                tags=tags,
                created_by=request.user,
                status='available',
                verification_status='pending'
            )
            
            # Handle file uploads
            if request.FILES.get('primary_photo'):
                collateral.primary_photo = request.FILES['primary_photo']
            if request.FILES.get('certificate_of_ownership'):
                collateral.certificate_of_ownership = request.FILES['certificate_of_ownership']
            if request.FILES.get('valuation_certificate'):
                collateral.valuation_certificate = request.FILES['valuation_certificate']
            if request.FILES.get('insurance_certificate'):
                collateral.insurance_certificate = request.FILES['insurance_certificate']
            
            collateral.save()
            
            messages.success(request, f'Collateral "{collateral.title}" created successfully!')
            return redirect('collateral:detail', pk=collateral.id)
            
        except Exception as e:
            messages.error(request, f'Error creating collateral: {str(e)}')
            return redirect('collateral:create')
    
    context = {
        'title': 'Add New Collateral',
        'clients': clients,
        'asset_types': asset_types,
    }
    
    return render(request, 'collateral/collateral_form.html', context)


@login_required
@user_passes_test(can_manage_collateral)
def collateral_detail(request, pk):
    """View collateral details"""
    collateral = get_object_or_404(Collateral, id=pk)
    
    # Get related data
    inspections = CollateralInspection.objects.filter(collateral=collateral).order_by('-inspection_date')
    movements = CollateralMovement.objects.filter(collateral=collateral).order_by('-movement_date')
    
    context = {
        'collateral': collateral,
        'inspections': inspections,
        'movements': movements,
    }
    
    return render(request, 'collateral/collateral_detail.html', context)


@login_required
@user_passes_test(can_manage_collateral)
def collateral_edit(request, pk):
    """Edit collateral"""
    collateral = get_object_or_404(Collateral, id=pk)
    
    clients = Client.objects.filter(status='active') if hasattr(Client, 'is_active') else Client.objects.all()
    asset_types = AssetType.objects.filter(is_active=True)
    
    if request.method == 'POST':
        try:
            # Update collateral fields
            collateral.title = request.POST.get('title', collateral.title)
            collateral.description = request.POST.get('description', collateral.description)
            collateral.make = request.POST.get('make', collateral.make)
            collateral.model = request.POST.get('model', collateral.model)
            collateral.year = request.POST.get('year') or None
            collateral.color = request.POST.get('color', collateral.color)
            collateral.condition = request.POST.get('condition', collateral.condition)
            collateral.estimated_value = request.POST.get('estimated_value', collateral.estimated_value)
            collateral.storage_location = request.POST.get('storage_location', collateral.storage_location)
            collateral.storage_section = request.POST.get('storage_section', collateral.storage_section)
            collateral.storage_shelf = request.POST.get('storage_shelf', collateral.storage_shelf)
            collateral.notes = request.POST.get('notes', collateral.notes)
            collateral.tags = request.POST.get('tags', collateral.tags)
            
            collateral.save()
            
            messages.success(request, f'Collateral "{collateral.title}" updated successfully!')
            return redirect('collateral:detail', pk=collateral.id)
            
        except Exception as e:
            messages.error(request, f'Error updating collateral: {str(e)}')
    
    context = {
        'collateral': collateral,
        'title': 'Edit Collateral',
        'clients': clients,
        'asset_types': asset_types,
    }
    
    return render(request, 'collateral/collateral_form.html', context)


@login_required
@user_passes_test(is_admin_or_ceo)
def collateral_delete(request, pk):
    """Delete collateral - Admin/CEO only"""
    collateral = get_object_or_404(Collateral, id=pk)
    
    if request.method == 'POST':
        title = collateral.title
        collateral.delete()
        messages.success(request, f'Collateral "{title}" deleted successfully!')
        return redirect('collateral:list')
    
    return render(request, 'collateral/collateral_confirm_delete.html', {'collateral': collateral})


@login_required
@user_passes_test(is_admin_or_ceo)
def asset_types(request):
    """View asset types - Admin/CEO only"""
    asset_types = AssetType.objects.all()
    return render(request, 'collateral/asset_types.html', {'asset_types': asset_types})


@login_required
@user_passes_test(is_admin_or_ceo)
def asset_type_create(request):
    """Create asset type - Admin/CEO only"""
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        icon = request.POST.get('icon', '')
        
        if name:
            AssetType.objects.create(name=name, description=description, icon=icon)
            messages.success(request, f'Asset type "{name}" created successfully!')
        else:
            messages.error(request, 'Name is required.')
        
        return redirect('collateral:asset_types')
    
    return render(request, 'collateral/asset_type_form.html')


@login_required
@user_passes_test(can_manage_collateral)
def inspections(request):
    """View inspections"""
    inspections = CollateralInspection.objects.all().select_related('collateral', 'inspected_by')
    return render(request, 'collateral/inspections.html', {'inspections': inspections})


@login_required
@user_passes_test(can_manage_collateral)
def inspection_create(request):
    """Create inspection"""
    if request.method == 'POST':
        try:
            inspection = CollateralInspection.objects.create(
                collateral_id=request.POST.get('collateral'),
                inspection_type=request.POST.get('inspection_type'),
                inspection_date=request.POST.get('inspection_date'),
                inspected_by=request.user,
                condition=request.POST.get('condition'),
                notes=request.POST.get('notes'),
                recommendation=request.POST.get('recommendation'),
                requires_action=request.POST.get('requires_action') == 'on'
            )
            
            if request.FILES.get('photos'):
                inspection.photos = request.FILES['photos']
                inspection.save()
            
            messages.success(request, 'Inspection recorded successfully!')
            return redirect('collateral:inspections')
            
        except Exception as e:
            messages.error(request, f'Error creating inspection: {str(e)}')
    
    collaterals = Collateral.objects.filter(status__in=['available', 'pledged'])
    
    context = {
        'collaterals': collaterals,
        'title': 'New Inspection'
    }
    return render(request, 'collateral/inspection_form.html', context)


@login_required
@user_passes_test(can_manage_collateral)
def inspection_detail(request, pk):
    """View inspection details"""
    inspection = get_object_or_404(CollateralInspection, id=pk)
    return render(request, 'collateral/inspection_detail.html', {'inspection': inspection})


@login_required
@user_passes_test(can_manage_collateral)
def movements(request):
    """View movements"""
    movements = CollateralMovement.objects.all().select_related('collateral', 'moved_by')
    return render(request, 'collateral/movements.html', {'movements': movements})


@login_required
@user_passes_test(can_manage_collateral)
def collateral_reports(request):
    """Collateral reports"""
    return render(request, 'collateral/reports.html')


@login_required
@user_passes_test(can_manage_collateral)
def verify_collateral(request, pk):
    """Verify collateral"""
    collateral = get_object_or_404(Collateral, id=pk)
    collateral.verify(request.user)
    messages.success(request, f'Collateral {collateral.collateral_id} verified successfully!')
    return redirect('collateral:detail', pk=collateral.id)


# ==================== INSURANCE VIEWS ====================

@login_required
@user_passes_test(can_manage_insurance)
def insurance_list(request):
    """List all insurance policies"""
    insured_collateral = Collateral.objects.filter(is_insured=True).select_related('client')
    
    # Calculate stats
    insured_count = insured_collateral.count()
    active_policies = insured_collateral.filter(
        insurance_expiry_date__gte=timezone.now().date()
    ).count()
    expiring_soon = insured_collateral.filter(
        insurance_expiry_date__gte=timezone.now().date(),
        insurance_expiry_date__lte=timezone.now().date() + timezone.timedelta(days=30)
    ).count()
    
    context = {
        'insured_collateral': insured_collateral,
        'insured_count': insured_count,
        'active_policies': active_policies,
        'expiring_soon': expiring_soon,
    }
    return render(request, 'collateral/insurance_list.html', context)


@login_required
@user_passes_test(can_manage_insurance)
def insurance_create(request):
    """Create or edit insurance policy"""
    collateral_id = request.GET.get('collateral_id') or request.POST.get('collateral_id')
    collateral = None
    
    if collateral_id:
        collateral = get_object_or_404(Collateral, id=collateral_id)
    
    if request.method == 'POST':
        collateral_id = request.POST.get('collateral_id')
        collateral = get_object_or_404(Collateral, id=collateral_id)
        
        collateral.is_insured = True
        collateral.insurance_provider = request.POST.get('insurance_provider')
        collateral.insurance_policy_number = request.POST.get('insurance_policy_number')
        collateral.insurance_value = request.POST.get('insurance_value')
        collateral.insurance_expiry_date = request.POST.get('insurance_expiry_date')
        collateral.save()
        
        messages.success(request, f'Insurance policy added for {collateral.title}')
        return redirect('collateral:insurance_list')
    
    collaterals = Collateral.objects.filter(is_insured=False).select_related('client')
    
    context = {
        'collaterals': collaterals,
        'selected_collateral': collateral,
        'title': 'Add Insurance Policy'
    }
    return render(request, 'collateral/insurance_form.html', context)


@login_required
@user_passes_test(can_manage_insurance)
def insurance_detail(request, pk):
    """View insurance details"""
    collateral = get_object_or_404(Collateral, id=pk, is_insured=True)
    return render(request, 'collateral/insurance_detail.html', {'collateral': collateral})


@login_required
@user_passes_test(can_manage_insurance)
def insurance_edit(request, pk):
    """Edit insurance policy"""
    collateral = get_object_or_404(Collateral, id=pk)
    
    if request.method == 'POST':
        collateral.insurance_provider = request.POST.get('insurance_provider')
        collateral.insurance_policy_number = request.POST.get('insurance_policy_number')
        collateral.insurance_value = request.POST.get('insurance_value')
        collateral.insurance_expiry_date = request.POST.get('insurance_expiry_date')
        collateral.is_insured = True
        collateral.save()
        
        messages.success(request, f'Insurance policy updated for {collateral.title}')
        return redirect('collateral:insurance_detail', pk=collateral.id)
    
    return render(request, 'collateral/insurance_form.html', {'collateral': collateral, 'title': 'Edit Insurance'})


@login_required
@user_passes_test(can_manage_insurance)
def insurance_delete(request, pk):
    """Delete insurance policy"""
    collateral = get_object_or_404(Collateral, id=pk)
    
    if request.method == 'POST':
        collateral.is_insured = False
        collateral.insurance_provider = None
        collateral.insurance_policy_number = None
        collateral.insurance_value = None
        collateral.insurance_expiry_date = None
        collateral.save()
        
        messages.success(request, f'Insurance policy removed from {collateral.title}')
        return redirect('collateral:insurance_list')
    
    return render(request, 'collateral/insurance_confirm_delete.html', {'collateral': collateral})


@login_required
@user_passes_test(can_manage_insurance)
def insurance_claims(request):
    """View insurance claims"""
    return render(request, 'collateral/insurance_claims.html', {'title': 'Insurance Claims'})


@login_required
@user_passes_test(can_manage_insurance)
def insurance_reports(request):
    """Insurance reports"""
    return render(request, 'collateral/insurance_reports.html', {'title': 'Insurance Reports'})


# ==================== API ENDPOINTS ====================

@login_required
def api_collateral_stats(request):
    """API endpoint for collateral statistics"""
    stats = {
        'total': Collateral.objects.count(),
        'pledged': Collateral.objects.filter(status='pledged').count(),
        'seized': Collateral.objects.filter(status='seized').count(),
        'verified': Collateral.objects.filter(verification_status='verified').count(),
        'pending': Collateral.objects.filter(verification_status='pending').count(),
        'total_value': float(Collateral.objects.aggregate(total=Sum('estimated_value'))['total'] or 0),
    }
    return JsonResponse(stats)


@login_required
def api_collateral_list(request):
    """API endpoint for collateral list"""
    collaterals = Collateral.objects.all().values(
        'id', 'collateral_id', 'title', 'serial_number', 'estimated_value',
        'condition', 'status', 'verification_status', 'client__first_name', 'client__last_name'
    )
    
    data = []
    for collateral in collaterals:
        data.append({
            'id': collateral['id'],
            'collateral_id': collateral['collateral_id'],
            'title': collateral['title'],
            'serial_number': collateral['serial_number'],
            'estimated_value': float(collateral['estimated_value']),
            'condition': collateral['condition'],
            'status': collateral['status'],
            'verification_status': collateral['verification_status'],
            'client_name': f"{collateral['client__first_name']} {collateral['client__last_name']}",
        })
    
    return JsonResponse({'collaterals': data})


@login_required
def api_insurance_stats(request):
    """API endpoint for insurance statistics"""
    insured_collateral = Collateral.objects.filter(is_insured=True)
    
    stats = {
        'insured_count': insured_collateral.count(),
        'active_policies': insured_collateral.filter(
            insurance_expiry_date__gte=timezone.now().date()
        ).count(),
        'expiring_soon': insured_collateral.filter(
            insurance_expiry_date__gte=timezone.now().date(),
            insurance_expiry_date__lte=timezone.now().date() + timezone.timedelta(days=30)
        ).count(),
    }
    return JsonResponse(stats)


# ==================== EXPORT FUNCTIONS ====================

@login_required
@user_passes_test(can_view_reports)
def export_collateral_csv(request):
    """Export collateral data to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="collateral_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Title', 'Serial Number', 'Client', 'Value', 'Condition', 'Status', 'Verification', 'Created'])
    
    collaterals = Collateral.objects.all().select_related('client')
    for collateral in collaterals:
        writer.writerow([
            collateral.collateral_id,
            collateral.title,
            collateral.serial_number,
            collateral.client.full_name,
            collateral.estimated_value,
            collateral.get_condition_display(),
            collateral.get_status_display(),
            collateral.get_verification_status_display(),
            collateral.created_at.strftime('%Y-%m-%d'),
        ])
    
    return response

# Add these functions to collateral/views.py

@login_required
@user_passes_test(is_admin_or_ceo)
def asset_type_edit(request, pk):
    """Edit asset type - Admin/CEO only"""
    asset_type = get_object_or_404(AssetType, id=pk)
    
    if request.method == 'POST':
        asset_type.name = request.POST.get('name', asset_type.name)
        asset_type.description = request.POST.get('description', asset_type.description)
        asset_type.icon = request.POST.get('icon', asset_type.icon)
        asset_type.save()
        messages.success(request, f'Asset type "{asset_type.name}" updated successfully!')
        return redirect('collateral:asset_types')
    
    return render(request, 'collateral/asset_type_form.html', {'asset_type': asset_type})


@login_required
@user_passes_test(is_admin_or_ceo)
def asset_type_delete(request, pk):
    """Delete asset type - Admin/CEO only"""
    asset_type = get_object_or_404(AssetType, id=pk)
    name = asset_type.name
    asset_type.delete()
    messages.success(request, f'Asset type "{name}" deleted successfully!')
    return redirect('collateral:asset_types')


@login_required
@user_passes_test(can_manage_collateral)
def inspection_edit(request, pk):
    """Edit inspection"""
    inspection = get_object_or_404(CollateralInspection, id=pk)
    
    if request.method == 'POST':
        try:
            inspection.inspection_type = request.POST.get('inspection_type', inspection.inspection_type)
            inspection.inspection_date = request.POST.get('inspection_date', inspection.inspection_date)
            inspection.condition = request.POST.get('condition', inspection.condition)
            inspection.notes = request.POST.get('notes', inspection.notes)
            inspection.recommendation = request.POST.get('recommendation', inspection.recommendation)
            inspection.requires_action = request.POST.get('requires_action') == 'on'
            
            if request.FILES.get('photos'):
                inspection.photos = request.FILES['photos']
            
            inspection.save()
            messages.success(request, 'Inspection updated successfully!')
            return redirect('collateral:inspection_detail', pk=inspection.id)
            
        except Exception as e:
            messages.error(request, f'Error updating inspection: {str(e)}')
    
    collaterals = Collateral.objects.filter(status__in=['available', 'pledged'])
    
    context = {
        'inspection': inspection,
        'collaterals': collaterals,
        'title': 'Edit Inspection'
    }
    return render(request, 'collateral/inspection_form.html', context)


@login_required
@user_passes_test(can_manage_collateral)
def movement_create(request):
    """Create movement record"""
    if request.method == 'POST':
        try:
            movement = CollateralMovement.objects.create(
                collateral_id=request.POST.get('collateral'),
                movement_type=request.POST.get('movement_type'),
                from_location=request.POST.get('from_location', ''),
                to_location=request.POST.get('to_location'),
                reason=request.POST.get('reason'),
                moved_by=request.user,
                notes=request.POST.get('notes', '')
            )
            messages.success(request, 'Movement recorded successfully!')
            return redirect('collateral:movements')
            
        except Exception as e:
            messages.error(request, f'Error creating movement: {str(e)}')
    
    collaterals = Collateral.objects.all()
    
    context = {
        'collaterals': collaterals,
        'title': 'New Movement'
    }
    return render(request, 'collateral/movement_form.html', context)


@login_required
@user_passes_test(can_manage_collateral)
def movement_detail(request, pk):
    """View movement details"""
    movement = get_object_or_404(CollateralMovement, id=pk)
    return render(request, 'collateral/movement_detail.html', {'movement': movement})


@login_required
@user_passes_test(can_view_reports)
def valuation_report(request):
    """Valuation report"""
    return render(request, 'collateral/valuation_report.html')


@login_required
@user_passes_test(can_view_reports)
def insurance_report(request):
    """Insurance report"""
    return render(request, 'collateral/insurance_report.html')


@login_required
@user_passes_test(can_manage_collateral)
def seize_collateral(request, pk):
    """Seize collateral"""
    collateral = get_object_or_404(Collateral, id=pk)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '')
        collateral.seize(request.user, reason)
        messages.success(request, f'Collateral {collateral.collateral_id} seized successfully!')
        return redirect('collateral:detail', pk=collateral.id)
    
    return render(request, 'collateral/seize_confirm.html', {'collateral': collateral})


@login_required
@user_passes_test(can_manage_collateral)
def release_collateral(request, pk):
    """Release collateral"""
    collateral = get_object_or_404(Collateral, id=pk)
    collateral.release(request.user)
    messages.success(request, f'Collateral {collateral.collateral_id} released successfully!')
    return redirect('collateral:detail', pk=collateral.id)


@login_required
@user_passes_test(can_manage_collateral)
def generate_qr_code(request, pk):
    """Generate QR code for collateral"""
    collateral = get_object_or_404(Collateral, id=pk)
    collateral.generate_qr_code()
    collateral.save()
    messages.success(request, f'QR code generated for {collateral.collateral_id}!')
    return redirect('collateral:detail', pk=collateral.id)


@login_required
def api_collateral_detail(request, pk):
    """API endpoint for collateral detail"""
    collateral = get_object_or_404(Collateral, id=pk)
    
    data = {
        'id': collateral.id,
        'collateral_id': collateral.collateral_id,
        'title': collateral.title,
        'description': collateral.description,
        'serial_number': collateral.serial_number,
        'estimated_value': float(collateral.estimated_value),
        'condition': collateral.condition,
        'status': collateral.status,
        'verification_status': collateral.verification_status,
        'client_name': collateral.client.full_name,
        'client_id': collateral.client.id,
        'created_at': collateral.created_at.isoformat() if collateral.created_at else None,
        'storage_location': collateral.storage_location,
    }
    
    return JsonResponse(data)


@login_required
@user_passes_test(can_view_reports)
def export_collateral_excel(request):
    """Export collateral data to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Collateral Report"
        
        # Headers
        headers = ['ID', 'Title', 'Serial Number', 'Client', 'Value (ZMW)', 'Condition', 'Status', 'Verification', 'Created']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        collaterals = Collateral.objects.all().select_related('client')
        for collateral in collaterals:
            ws.append([
                collateral.collateral_id,
                collateral.title,
                collateral.serial_number,
                collateral.client.full_name,
                float(collateral.estimated_value),
                collateral.get_condition_display(),
                collateral.get_status_display(),
                collateral.get_verification_status_display(),
                collateral.created_at.strftime('%Y-%m-%d'),
            ])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="collateral_export.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'openpyxl is not installed.')
        return redirect('collateral:list')
    
# collateral/views.py - Add these improved report views at the end of the file

# ==================== COMPREHENSIVE REPORTS SECTION ====================

@login_required
@user_passes_test(can_view_reports)
def collateral_reports_dashboard(request):
    """Main collateral reports dashboard"""
    from django.db.models import Sum, Count, Q
    from django.utils import timezone
    from datetime import timedelta
    
    # Get date range filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    report_type = request.GET.get('report_type', 'summary')
    
    # Base queryset
    collaterals = Collateral.objects.all()
    
    # Apply date filters
    if date_from:
        collaterals = collaterals.filter(created_at__date__gte=date_from)
    if date_to:
        collaterals = collaterals.filter(created_at__date__lte=date_to)
    
    # Summary Statistics
    total_collateral = collaterals.count()
    total_value = collaterals.aggregate(total=Sum('estimated_value'))['total'] or 0
    avg_value = total_value / total_collateral if total_collateral > 0 else 0
    
    # Status breakdown
    status_stats = {
        'available': collaterals.filter(status='available').count(),
        'pledged': collaterals.filter(status='pledged').count(),
        'seized': collaterals.filter(status='seized').count(),
        'sold': collaterals.filter(status='sold').count(),
        'released': collaterals.filter(status='released').count(),
    }
    
    # Verification status
    verification_stats = {
        'pending': collaterals.filter(verification_status='pending').count(),
        'verified': collaterals.filter(verification_status='verified').count(),
        'rejected': collaterals.filter(verification_status='rejected').count(),
    }
    
    # Condition breakdown
    condition_stats = {
        'excellent': collaterals.filter(condition='excellent').count(),
        'good': collaterals.filter(condition='good').count(),
        'fair': collaterals.filter(condition='fair').count(),
        'poor': collaterals.filter(condition='poor').count(),
        'damaged': collaterals.filter(condition='damaged').count(),
    }
    
    # Asset type breakdown
    asset_type_stats = []
    asset_types = AssetType.objects.filter(is_active=True)
    for asset_type in asset_types:
        count = collaterals.filter(asset_type=asset_type).count()
        value = collaterals.filter(asset_type=asset_type).aggregate(total=Sum('estimated_value'))['total'] or 0
        if count > 0:
            asset_type_stats.append({
                'name': asset_type.name,
                'icon': asset_type.icon,
                'count': count,
                'value': float(value),
                'percentage': (count / total_collateral * 100) if total_collateral > 0 else 0
            })
    
    # Monthly trend (last 12 months)
    monthly_data = []
    today = timezone.now().date()
    for i in range(11, -1, -1):
        month_date = today - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year+1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month_date.replace(month=month_date.month+1, day=1) - timedelta(days=1)
        
        month_collaterals = Collateral.objects.filter(
            created_at__date__gte=month_start,
            created_at__date__lte=month_end
        )
        
        monthly_data.append({
            'month': month_start.strftime('%b %Y'),
            'count': month_collaterals.count(),
            'value': float(month_collaterals.aggregate(total=Sum('estimated_value'))['total'] or 0)
        })
    
    # Insurance statistics
    insured_collaterals = collaterals.filter(is_insured=True)
    insurance_stats = {
        'insured_count': insured_collaterals.count(),
        'insured_value': float(insured_collaterals.aggregate(total=Sum('insurance_value'))['total'] or 0),
        'active_policies': insured_collaterals.filter(
            insurance_expiry_date__gte=timezone.now().date()
        ).count(),
        'expiring_soon': insured_collaterals.filter(
            insurance_expiry_date__gte=timezone.now().date(),
            insurance_expiry_date__lte=timezone.now().date() + timedelta(days=30)
        ).count(),
        'expired': insured_collaterals.filter(
            insurance_expiry_date__lt=timezone.now().date()
        ).count(),
    }
    
    # Recent collaterals
    recent_collaterals = collaterals.select_related('client', 'asset_type').order_by('-created_at')[:10]
    
    # Top clients by collateral value
    top_clients = []
    client_collaterals = collaterals.values('client__id', 'client__full_name').annotate(
        total_value=Sum('estimated_value'),
        count=Count('id')
    ).order_by('-total_value')[:10]
    
    for client in client_collaterals:
        top_clients.append({
            'name': client['client__full_name'],
            'count': client['count'],
            'value': float(client['total_value'])
        })
    
    context = {
        # Summary
        'report_title': 'Collateral Reports Dashboard',
        'report_date': timezone.now(),
        'date_from': date_from,
        'date_to': date_to,
        'report_type': report_type,
        
        # Statistics
        'total_collateral': total_collateral,
        'total_value': float(total_value),
        'avg_value': float(avg_value),
        
        # Breakdowns
        'status_stats': status_stats,
        'verification_stats': verification_stats,
        'condition_stats': condition_stats,
        'asset_type_stats': asset_type_stats,
        
        # Trends
        'monthly_data': monthly_data,
        
        # Insurance
        'insurance_stats': insurance_stats,
        
        # Lists
        'recent_collaterals': recent_collaterals,
        'top_clients': top_clients,
    }
    
    # Check for print format
    if request.GET.get('format') == 'print':
        return render(request, 'collateral/reports_print.html', context)
    
    return render(request, 'collateral/reports_dashboard.html', context)


@login_required
@user_passes_test(can_view_reports)
def valuation_report(request):
    """Valuation report - detailed asset valuation"""
    from django.db.models import Sum, Avg, Q
    
    # Get filters
    asset_type_filter = request.GET.get('asset_type', '')
    min_value = request.GET.get('min_value', '')
    max_value = request.GET.get('max_value', '')
    
    # Base queryset
    collaterals = Collateral.objects.all().select_related('client', 'asset_type')
    
    # Apply filters
    if asset_type_filter:
        collaterals = collaterals.filter(asset_type__name=asset_type_filter)
    if min_value:
        collaterals = collaterals.filter(estimated_value__gte=min_value)
    if max_value:
        collaterals = collaterals.filter(estimated_value__lte=max_value)
    
    # Statistics
    total_collateral = collaterals.count()
    total_value = collaterals.aggregate(total=Sum('estimated_value'))['total'] or 0
    avg_value = collaterals.aggregate(avg=Avg('estimated_value'))['avg'] or 0
    max_asset_value = collaterals.aggregate(max=Sum('estimated_value'))['max'] or 0
    min_asset_value = collaterals.aggregate(min=Sum('estimated_value'))['min'] or 0
    
    # Value distribution
    value_ranges = [
        {'range': '0 - 1,000', 'min': 0, 'max': 1000, 'count': 0, 'value': 0},
        {'range': '1,001 - 5,000', 'min': 1001, 'max': 5000, 'count': 0, 'value': 0},
        {'range': '5,001 - 10,000', 'min': 5001, 'max': 10000, 'count': 0, 'value': 0},
        {'range': '10,001 - 25,000', 'min': 10001, 'max': 25000, 'count': 0, 'value': 0},
        {'range': '25,001 - 50,000', 'min': 25001, 'max': 50000, 'count': 0, 'value': 0},
        {'range': '50,000+', 'min': 50001, 'max': 999999999, 'count': 0, 'value': 0},
    ]
    
    for collateral in collaterals:
        value = float(collateral.estimated_value)
        for range_data in value_ranges:
            if range_data['min'] <= value <= range_data['max']:
                range_data['count'] += 1
                range_data['value'] += value
                break
    
    # Collateral by asset type with values
    asset_type_valuation = []
    asset_types = AssetType.objects.filter(is_active=True)
    for asset_type in asset_types:
        type_collaterals = collaterals.filter(asset_type=asset_type)
        count = type_collaterals.count()
        if count > 0:
            asset_type_valuation.append({
                'name': asset_type.name,
                'count': count,
                'total_value': float(type_collaterals.aggregate(total=Sum('estimated_value'))['total'] or 0),
                'avg_value': float(type_collaterals.aggregate(avg=Avg('estimated_value'))['avg'] or 0),
            })
    
    # All collaterals with values
    all_collaterals = collaterals.order_by('-estimated_value')[:50]
    
    context = {
        'title': 'Valuation Report',
        'collaterals': all_collaterals,
        'total_collateral': total_collateral,
        'total_value': float(total_value),
        'avg_value': float(avg_value),
        'max_asset_value': float(max_asset_value),
        'min_asset_value': float(min_asset_value),
        'value_ranges': value_ranges,
        'asset_type_valuation': asset_type_valuation,
        'asset_type_filter': asset_type_filter,
        'min_value': min_value,
        'max_value': max_value,
        'report_date': timezone.now(),
    }
    
    if request.GET.get('format') == 'print':
        return render(request, 'collateral/valuation_report_print.html', context)
    
    return render(request, 'collateral/valuation_report.html', context)


@login_required
@user_passes_test(can_view_reports)
def insurance_report(request):
    """Insurance report - detailed insurance analysis"""
    from django.db.models import Sum, Count, Q
    from django.utils import timezone
    from datetime import timedelta
    
    # Get filters
    status_filter = request.GET.get('status', 'all')
    provider_filter = request.GET.get('provider', '')
    
    # Base queryset - only insured collaterals
    collaterals = Collateral.objects.filter(is_insured=True).select_related('client', 'asset_type')
    
    # Apply filters
    today = timezone.now().date()
    if status_filter == 'active':
        collaterals = collaterals.filter(insurance_expiry_date__gte=today)
    elif status_filter == 'expiring':
        collaterals = collaterals.filter(
            insurance_expiry_date__gte=today,
            insurance_expiry_date__lte=today + timedelta(days=30)
        )
    elif status_filter == 'expired':
        collaterals = collaterals.filter(insurance_expiry_date__lt=today)
    
    if provider_filter:
        collaterals = collaterals.filter(insurance_provider__icontains=provider_filter)
    
    # Statistics
    total_insured = collaterals.count()
    total_insured_value = collaterals.aggregate(total=Sum('insurance_value'))['total'] or 0
    total_asset_value = collaterals.aggregate(total=Sum('estimated_value'))['total'] or 0
    
    # Status breakdown
    active_policies = collaterals.filter(insurance_expiry_date__gte=today).count()
    expiring_soon = collaterals.filter(
        insurance_expiry_date__gte=today,
        insurance_expiry_date__lte=today + timedelta(days=30)
    ).count()
    expired_policies = collaterals.filter(insurance_expiry_date__lt=today).count()
    
    # Provider breakdown
    provider_stats = []
    providers = collaterals.values('insurance_provider').annotate(
        count=Count('id'),
        total_value=Sum('insurance_value'),
        asset_value=Sum('estimated_value')
    ).order_by('-total_value')
    
    for provider in providers:
        if provider['insurance_provider']:
            provider_stats.append({
                'name': provider['insurance_provider'],
                'count': provider['count'],
                'total_value': float(provider['total_value'] or 0),
                'asset_value': float(provider['asset_value'] or 0),
            })
    
    # Monthly expiry forecast
    expiry_forecast = []
    for i in range(12):
        future_date = today + timedelta(days=30 * i)
        month_start = future_date.replace(day=1)
        if future_date.month == 12:
            month_end = future_date.replace(year=future_date.year+1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = future_date.replace(month=future_date.month+1, day=1) - timedelta(days=1)
        
        expiring = collaterals.filter(
            insurance_expiry_date__gte=month_start,
            insurance_expiry_date__lte=month_end
        ).count()
        
        expiry_forecast.append({
            'month': month_start.strftime('%b %Y'),
            'expiring_count': expiring
        })
    
    # List of insured collaterals
    insured_list = collaterals.order_by('insurance_expiry_date')[:50]
    
    context = {
        'title': 'Insurance Report',
        'collaterals': insured_list,
        'total_insured': total_insured,
        'total_insured_value': float(total_insured_value),
        'total_asset_value': float(total_asset_value),
        'coverage_percentage': (total_insured_value / total_asset_value * 100) if total_asset_value > 0 else 0,
        'active_policies': active_policies,
        'expiring_soon': expiring_soon,
        'expired_policies': expired_policies,
        'provider_stats': provider_stats,
        'expiry_forecast': expiry_forecast,
        'status_filter': status_filter,
        'provider_filter': provider_filter,
        'report_date': timezone.now(),
    }
    
    if request.GET.get('format') == 'print':
        return render(request, 'collateral/insurance_report_print.html', context)
    
    return render(request, 'collateral/insurance_report.html', context)


@login_required
@user_passes_test(can_view_reports)
def movement_report(request):
    """Movement report - track collateral movements"""
    from django.db.models import Count, Q
    from datetime import timedelta
    
    # Get filters
    movement_type = request.GET.get('movement_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset
    movements = CollateralMovement.objects.all().select_related('collateral', 'moved_by')
    
    # Apply filters
    if movement_type:
        movements = movements.filter(movement_type=movement_type)
    if date_from:
        movements = movements.filter(movement_date__date__gte=date_from)
    if date_to:
        movements = movements.filter(movement_date__date__lte=date_to)
    
    # Statistics
    total_movements = movements.count()
    
    # Movement type breakdown
    type_stats = []
    movement_types = [
        ('inbound', 'Inbound'),
        ('outbound', 'Outbound'),
        ('transfer', 'Transfer'),
        ('loan', 'Loan'),
        ('return', 'Return'),
    ]
    
    for m_type, m_label in movement_types:
        count = movements.filter(movement_type=m_type).count()
        type_stats.append({
            'type': m_label,
            'count': count,
            'percentage': (count / total_movements * 100) if total_movements > 0 else 0
        })
    
    # Recent movements
    recent_movements = movements.order_by('-movement_date')[:50]
    
    # Most moved collaterals
    most_moved = movements.values('collateral__title', 'collateral__collateral_id').annotate(
        move_count=Count('id')
    ).order_by('-move_count')[:10]
    
    context = {
        'title': 'Movement Report',
        'movements': recent_movements,
        'total_movements': total_movements,
        'type_stats': type_stats,
        'most_moved': most_moved,
        'movement_type': movement_type,
        'date_from': date_from,
        'date_to': date_to,
        'report_date': timezone.now(),
    }
    
    if request.GET.get('format') == 'print':
        return render(request, 'collateral/movement_report_print.html', context)
    
    return render(request, 'collateral/movement_report.html', context)


@login_required
@user_passes_test(can_view_reports)
def inspection_report(request):
    """Inspection report - track inspections and findings"""
    from django.db.models import Count, Q
    from datetime import timedelta
    
    # Get filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    requires_action = request.GET.get('requires_action', '')
    
    # Base queryset
    inspections = CollateralInspection.objects.all().select_related('collateral', 'inspected_by')
    
    # Apply filters
    if date_from:
        inspections = inspections.filter(inspection_date__gte=date_from)
    if date_to:
        inspections = inspections.filter(inspection_date__lte=date_to)
    if requires_action == 'yes':
        inspections = inspections.filter(requires_action=True)
    
    # Statistics
    total_inspections = inspections.count()
    
    # Inspection type breakdown
    type_stats = []
    inspection_types = [
        ('initial', 'Initial'),
        ('quarterly', 'Quarterly'),
        ('annual', 'Annual'),
        ('special', 'Special'),
    ]
    
    for i_type, i_label in inspection_types:
        count = inspections.filter(inspection_type=i_type).count()
        type_stats.append({
            'type': i_label,
            'count': count,
            'percentage': (count / total_inspections * 100) if total_inspections > 0 else 0
        })
    
    # Condition breakdown
    condition_stats = []
    conditions = [
        ('excellent', 'Excellent'),
        ('good', 'Good'),
        ('fair', 'Fair'),
        ('poor', 'Poor'),
        ('damaged', 'Damaged'),
    ]
    
    for cond, cond_label in conditions:
        count = inspections.filter(condition=cond).count()
        condition_stats.append({
            'condition': cond_label,
            'count': count,
            'percentage': (count / total_inspections * 100) if total_inspections > 0 else 0
        })
    
    # Requires action
    pending_actions = inspections.filter(requires_action=True).count()
    
    # Recent inspections
    recent_inspections = inspections.order_by('-inspection_date')[:50]
    
    context = {
        'title': 'Inspection Report',
        'inspections': recent_inspections,
        'total_inspections': total_inspections,
        'pending_actions': pending_actions,
        'type_stats': type_stats,
        'condition_stats': condition_stats,
        'date_from': date_from,
        'date_to': date_to,
        'requires_action': requires_action,
        'report_date': timezone.now(),
    }
    
    if request.GET.get('format') == 'print':
        return render(request, 'collateral/inspection_report_print.html', context)
    
    return render(request, 'collateral/inspection_report.html', context)


# Update the collateral_reports function to redirect to the new dashboard
@login_required
@user_passes_test(can_view_reports)
def collateral_reports_legacy(request):
    """Legacy collateral reports - redirect to new dashboard"""
    from django.shortcuts import redirect
    return redirect('collateral:reports_dashboard')